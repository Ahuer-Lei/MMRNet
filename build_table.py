import os
import torch 
import json
import openpyxl
from openpyxl.styles import Font,Alignment
import torch.nn as nn
from datetime import datetime

address_dict={ "result_save_path": '/data/zcl3/CVPR2025/Major_revision/code/Our/Evaluator.xlsx'}

# 将实验结果输出为xlsx文件
def set_result(result,method_dict):
    '''
    将实验结果输出为xlsx文件
    method : 实验方法
    '''
    data_time=str(datetime.now())[0:19]
    experiment_result=[data_time,method_dict['method'],method_dict['id'],method_dict['batch_size'],method_dict['epoch']]
    
    if not os.path.exists(address_dict['result_save_path']):
        wb=openpyxl.Workbook()
        ws=wb.active
        ws.title='experiment_result'
        table_head=['日期','实验方法','实验编号','batch_size','epoch']
        for item in result.items():
            table_head.append(item[0])
            experiment_result.append(item[1])
        ws.append(table_head)
        ws.append(experiment_result)
    else:
        wb=openpyxl.load_workbook(address_dict['result_save_path'])
        ws=wb['experiment_result']
        for item in result.items():
            experiment_result.append(item[1])
        ws.append(experiment_result)
    
    wb.save(filename=address_dict['result_save_path']) 
    wb.close()
    set_font_style(address_dict['result_save_path'])


def set_font_style(path):
    '''
    调整excel字体和格式
    '''
    if os.path.exists(path):
        wb=openpyxl.load_workbook(path)
       
        ws=wb['experiment_result']
    
        font = Font(name="微软雅黑",size=10)
        align = Alignment(horizontal="center",vertical="center")
        font_=Font(bold=True)
        max_rows = ws.max_row  # 获取最大行
        max_columns = ws.max_column  # 获取最大列

        # openpyxl的下标从1开始
        for i in range(1, max_rows + 1):
            for j in range(1, max_columns + 1):
                ws.cell(i, j).alignment = align
                ws.cell(i,j).font=font
        if path[-11:-5]=='result':
            for i in range(2, max_rows + 1):
                for j in range(5, max_columns + 1):
                    ws.cell(i,j).number_format='0.0000'
            
        else:
            for i in range(2, max_rows + 1):
                for j in [6,7,8]:
                    ws.cell(i,j).number_format='0.0000'
            
        for j in range(1, max_columns + 1):
            ws.cell(1,j).font=font_

        wb.save(path)
        wb.close()


# if __name__=='__main__':
#         result= {'avg_pixel_error':0.16, 'detail':'使用预训练好的网络指导模态差异学习', 'loss':'loss_sim+10*loss_tp+100 *loss_guild'}
#         method_dict={'method':'Net','id':0,'batch_size':6,'epoch':152}
#         set_result(result=result,method_dict=method_dict)  
    

    